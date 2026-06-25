from __future__ import annotations

import argparse
import json
import math
import re
import struct
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TEXT_RE = re.compile(r'"(?:[^"]|"")*"')
FUNC_RE = re.compile(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9_.]*)\s*\(", re.I)
PROC_RE = re.compile(
    r"^\s*(?:Public\s+|Private\s+|Friend\s+|Static\s+)?"
    r"(Sub|Function|Property\s+(?:Get|Let|Set))\s+([A-Za-z_][A-Za-z0-9_]*)",
    re.I | re.M,
)
CALL_KEYWORDS = [
    "Application",
    "Workbook",
    "Worksheet",
    "Range",
    "Cells",
    "ListObject",
    "PivotTable",
    "UserForm",
    "MsgBox",
    "InputBox",
    "CreateObject",
    "Shell",
    "Kill",
    "FileSystemObject",
    "ADODB",
    "QueryTables",
    "On Error",
]
DIR_RECORD_MODULES = 0x000F
DIR_RECORD_MODULENAME = 0x0019
DIR_RECORD_MODULESTREAMNAME = 0x001A
DIR_RECORD_MODULEDOCSTRING = 0x001C
DIR_RECORD_MODULEOFFSET = 0x0031
DIR_RECORD_MODULETYPE_PROCEDURAL = 0x0021
DIR_RECORD_MODULETYPE_DOCUMENT = 0x0022
DIR_RECORD_TERMINATOR = 0x002B


def compact_formula(formula: Any) -> str | None:
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    sanitized = TEXT_RE.sub('"TEXT"', formula)
    sanitized = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "CELL", sanitized)
    sanitized = re.sub(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", "COL:COL", sanitized)
    sanitized = re.sub(r"\d+(?:\.\d+)?", "N", sanitized)
    return sanitized[:240]


def formula_functions(formula: str) -> list[str]:
    if not formula:
        return []
    return sorted({m.group(1).upper() for m in FUNC_RE.finditer(formula)})


def value_kind(cell: Any) -> str:
    value = cell.value
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if getattr(value, "year", None) and getattr(value, "month", None):
        return "date"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "text"


def likely_header_row(ws, scan_rows: int = 30) -> int | None:
    best_row = None
    best_score = -1
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        cells = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        non_empty = [v for v in cells if v not in (None, "")]
        text_count = sum(1 for v in non_empty if isinstance(v, str) and not str(v).startswith("="))
        score = text_count * 2 + len(non_empty)
        if score > best_score:
            best_row = row
            best_score = score
    return best_row if best_score > 0 else None


def sheet_profile(ws) -> dict[str, Any]:
    formula_counter: Counter[str] = Counter()
    formula_samples: dict[str, dict[str, Any]] = {}
    formula_functions_counter: Counter[str] = Counter()
    type_counter: Counter[str] = Counter()
    column_profiles: list[dict[str, Any]] = []
    header_row = likely_header_row(ws)

    for row in ws.iter_rows():
        for cell in row:
            kind = value_kind(cell)
            type_counter[kind] += 1
            if kind == "formula":
                skeleton = compact_formula(cell.value) or "=FORMULA"
                formula_counter[skeleton] += 1
                formula_functions_counter.update(formula_functions(str(cell.value)))
                formula_samples.setdefault(
                    skeleton,
                    {"first_cell": cell.coordinate, "functions": formula_functions(str(cell.value))},
                )

    for col in range(1, ws.max_column + 1):
        counts: Counter[str] = Counter()
        non_empty = 0
        formula_count = 0
        number_formats: Counter[str] = Counter()
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, col)
            kind = value_kind(cell)
            counts[kind] += 1
            if kind != "blank":
                non_empty += 1
            if kind == "formula":
                formula_count += 1
            if cell.number_format and cell.number_format != "General":
                number_formats[cell.number_format] += 1
        header = None
        if header_row:
            hv = ws.cell(header_row, col).value
            if hv not in (None, "") and isinstance(hv, str):
                header = hv[:80]
        if non_empty or formula_count:
            column_profiles.append(
                {
                    "column": get_column_letter(col),
                    "header": header,
                    "non_empty_count": non_empty,
                    "formula_count": formula_count,
                    "kinds": dict(counts),
                    "number_formats": dict(number_formats.most_common(3)),
                }
            )

    validations = []
    for dv in getattr(ws.data_validations, "dataValidation", []):
        validations.append(
            {
                "type": dv.type,
                "operator": dv.operator,
                "ranges": str(dv.sqref),
                "allow_blank": dv.allowBlank,
            }
        )

    conditional_formats = []
    try:
        cf_items = ws.conditional_formatting._cf_rules.items()
        for range_obj, rules in cf_items:
            conditional_formats.append(
                {
                    "range": str(range_obj),
                    "rules": [getattr(rule, "type", "unknown") for rule in rules],
                }
            )
    except Exception:
        conditional_formats = [{"count": len(ws.conditional_formatting)}]

    tables = []
    for table in ws.tables.values():
        tables.append({"name": table.name, "ref": table.ref, "style": getattr(table.tableStyleInfo, "name", None)})

    return {
        "name": ws.title,
        "dimensions": {"max_row": ws.max_row, "max_column": ws.max_column},
        "header_row_guess": header_row,
        "type_counts": dict(type_counter),
        "formula_count": sum(formula_counter.values()),
        "formula_function_counts": dict(formula_functions_counter.most_common()),
        "formula_patterns": [
            {
                "pattern": pattern,
                "count": count,
                "first_cell": formula_samples[pattern]["first_cell"],
                "functions": formula_samples[pattern]["functions"],
            }
            for pattern, count in formula_counter.most_common(25)
        ],
        "column_profiles": column_profiles,
        "tables": tables,
        "merged_ranges_count": len(ws.merged_cells.ranges),
        "merged_ranges_sample": [str(rng) for rng in list(ws.merged_cells.ranges)[:12]],
        "data_validations": validations[:50],
        "data_validation_count": len(validations),
        "conditional_formats": conditional_formats[:50],
        "conditional_format_count": len(conditional_formats),
        "charts_count": len(getattr(ws, "_charts", [])),
        "images_count": len(getattr(ws, "_images", [])),
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
    }


def xml_feature_profile(xlsm_path: Path) -> dict[str, Any]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    profile: dict[str, Any] = {"entries": [], "vba_project": False, "worksheets": []}
    with zipfile.ZipFile(xlsm_path) as zf:
        names = zf.namelist()
        profile["entries"] = sorted(names)
        profile["vba_project"] = "xl/vbaProject.bin" in names
        profile["tables_count"] = len([n for n in names if n.startswith("xl/tables/") and n.endswith(".xml")])
        profile["drawings_count"] = len([n for n in names if n.startswith("xl/drawings/") and n.endswith(".xml")])
        profile["media_count"] = len([n for n in names if n.startswith("xl/media/")])
        profile["control_properties_count"] = len([n for n in names if n.startswith("xl/ctrlProps/")])
        profile["active_x_count"] = len([n for n in names if n.startswith("xl/activeX/")])
        profile["custom_ui"] = [n for n in names if n.startswith("customUI/")]

        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = {}
        if "xl/_rels/workbook.xml.rels" in names:
            rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            for rel in rel_root:
                rels[rel.attrib.get("Id")] = rel.attrib

        sheet_id_to_name = {}
        sheet_order = []
        for sheet in workbook_xml.findall(".//main:sheet", ns):
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            name = sheet.attrib.get("name")
            target = rels.get(rid, {}).get("Target")
            if target:
                sheet_path = "xl/" + target.lstrip("/")
                sheet_id_to_name[sheet_path.replace("\\", "/")] = name
            sheet_order.append(name)
        profile["sheet_order"] = sheet_order

        for sheet_path in sorted(n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
            root = ET.fromstring(zf.read(sheet_path))
            def count(tag: str) -> int:
                return len(root.findall(f".//main:{tag}", ns))

            feature = {
                "path": sheet_path,
                "name": sheet_id_to_name.get(sheet_path),
                "dimension": root.find("main:dimension", ns).attrib.get("ref")
                if root.find("main:dimension", ns) is not None
                else None,
                "sheetViews": count("sheetView"),
                "mergeCells": count("mergeCell"),
                "conditionalFormatting": count("conditionalFormatting"),
                "dataValidations": count("dataValidation"),
                "hyperlinks": count("hyperlink"),
                "tableParts": count("tablePart"),
                "oleObjects": count("oleObject"),
                "controls": count("control"),
                "drawings": count("drawing"),
                "legacyDrawings": count("legacyDrawing"),
                "sheetProtection": count("sheetProtection"),
            }
            profile["worksheets"].append(feature)
    return profile


FREESECT = 0xFFFFFFFF
ENDOFCHAIN = 0xFFFFFFFE
FATSECT = 0xFFFFFFFD
DIFSECT = 0xFFFFFFFC
NOSTREAM = 0xFFFFFFFF


@dataclass
class DirEntry:
    id: int
    name: str
    type: int
    left: int
    right: int
    child: int
    start_sector: int
    stream_size: int


class CompoundFile:
    def __init__(self, data: bytes):
        self.data = data
        if data[:8] != bytes.fromhex("D0CF11E0A1B11AE1"):
            raise ValueError("Not a compound binary file")
        self.sector_size = 1 << struct.unpack_from("<H", data, 30)[0]
        self.mini_sector_size = 1 << struct.unpack_from("<H", data, 32)[0]
        self.num_fat_sectors = struct.unpack_from("<I", data, 44)[0]
        self.first_dir_sector = struct.unpack_from("<I", data, 48)[0]
        self.mini_cutoff = struct.unpack_from("<I", data, 56)[0]
        self.first_mini_fat_sector = struct.unpack_from("<I", data, 60)[0]
        self.num_mini_fat_sectors = struct.unpack_from("<I", data, 64)[0]
        self.first_difat_sector = struct.unpack_from("<I", data, 68)[0]
        self.num_difat_sectors = struct.unpack_from("<I", data, 72)[0]
        self.difat = list(struct.unpack_from("<109I", data, 76))
        self._load_fat()
        self._load_directory()
        self._load_mini_fat()

    def sector(self, sid: int) -> bytes:
        start = 512 + sid * self.sector_size
        return self.data[start : start + self.sector_size]

    def chain(self, start_sid: int, fat: list[int] | None = None) -> list[int]:
        if start_sid in (FREESECT, ENDOFCHAIN, NOSTREAM):
            return []
        table = self.fat if fat is None else fat
        result = []
        sid = start_sid
        seen = set()
        while sid not in (FREESECT, ENDOFCHAIN, NOSTREAM) and sid not in seen:
            seen.add(sid)
            result.append(sid)
            if sid >= len(table):
                break
            sid = table[sid]
        return result

    def read_sector_chain(self, start_sid: int, size: int | None = None) -> bytes:
        blob = b"".join(self.sector(sid) for sid in self.chain(start_sid))
        return blob[:size] if size is not None else blob

    def _load_fat(self) -> None:
        difat = [sid for sid in self.difat if sid not in (FREESECT, ENDOFCHAIN)]
        next_difat = self.first_difat_sector
        for _ in range(self.num_difat_sectors):
            if next_difat in (FREESECT, ENDOFCHAIN):
                break
            sector = self.sector(next_difat)
            entries = list(struct.unpack_from(f"<{self.sector_size // 4}I", sector, 0))
            difat.extend(sid for sid in entries[:-1] if sid not in (FREESECT, ENDOFCHAIN))
            next_difat = entries[-1]
        fat_entries: list[int] = []
        for sid in difat[: self.num_fat_sectors]:
            sector = self.sector(sid)
            fat_entries.extend(struct.unpack_from(f"<{self.sector_size // 4}I", sector, 0))
        self.fat = fat_entries

    def _load_directory(self) -> None:
        raw = self.read_sector_chain(self.first_dir_sector)
        self.entries: list[DirEntry] = []
        for idx in range(0, len(raw), 128):
            entry = raw[idx : idx + 128]
            if len(entry) < 128:
                continue
            name_len = struct.unpack_from("<H", entry, 64)[0]
            name = entry[: max(0, name_len - 2)].decode("utf-16le", errors="ignore")
            entry_type = entry[66]
            if not name:
                continue
            stream_size = struct.unpack_from("<Q", entry, 120)[0]
            self.entries.append(
                DirEntry(
                    id=idx // 128,
                    name=name,
                    type=entry_type,
                    left=struct.unpack_from("<I", entry, 68)[0],
                    right=struct.unpack_from("<I", entry, 72)[0],
                    child=struct.unpack_from("<I", entry, 76)[0],
                    start_sector=struct.unpack_from("<I", entry, 116)[0],
                    stream_size=stream_size,
                )
            )
        self.root = next((e for e in self.entries if e.type == 5), self.entries[0])

    def _load_mini_fat(self) -> None:
        raw = self.read_sector_chain(self.first_mini_fat_sector) if self.num_mini_fat_sectors else b""
        self.mini_fat = list(struct.unpack_from(f"<{len(raw) // 4}I", raw, 0)) if raw else []
        self.mini_stream = self.read_sector_chain(self.root.start_sector, self.root.stream_size)

    def sibling_tree(self, sid: int) -> list[DirEntry]:
        result: list[DirEntry] = []
        if sid in (NOSTREAM, FREESECT) or sid >= len(self.entries):
            return result
        entry = self.entries[sid]
        result.extend(self.sibling_tree(entry.left))
        result.append(entry)
        result.extend(self.sibling_tree(entry.right))
        return result

    def list_paths(self) -> dict[str, DirEntry]:
        paths: dict[str, DirEntry] = {}

        def walk(storage: DirEntry, prefix: str) -> None:
            for child in self.sibling_tree(storage.child):
                path = f"{prefix}/{child.name}" if prefix else child.name
                paths[path] = child
                if child.type == 1:
                    walk(child, path)

        walk(self.root, "")
        return paths

    def stream(self, entry: DirEntry) -> bytes:
        if entry.stream_size < self.mini_cutoff and entry.type == 2:
            chunks = []
            sid = entry.start_sector
            seen = set()
            while sid not in (FREESECT, ENDOFCHAIN, NOSTREAM) and sid not in seen:
                seen.add(sid)
                start = sid * self.mini_sector_size
                chunks.append(self.mini_stream[start : start + self.mini_sector_size])
                if sid >= len(self.mini_fat):
                    break
                sid = self.mini_fat[sid]
            return b"".join(chunks)[: entry.stream_size]
        return self.read_sector_chain(entry.start_sector, entry.stream_size)


def decompress_vba(data: bytes) -> bytes:
    if not data or data[0] != 0x01:
        return data
    pos = 1
    out = bytearray()
    while pos + 2 <= len(data):
        header = struct.unpack_from("<H", data, pos)[0]
        pos += 2
        chunk_size = (header & 0x0FFF) + 3
        compressed = bool(header & 0x8000)
        chunk_end = min(pos + chunk_size - 2, len(data))
        chunk_start_out = len(out)
        if not compressed:
            out.extend(data[pos:chunk_end])
            pos = chunk_end
            continue
        while pos < chunk_end:
            flags = data[pos]
            pos += 1
            for bit in range(8):
                if pos >= chunk_end:
                    break
                if not (flags & (1 << bit)):
                    out.append(data[pos])
                    pos += 1
                else:
                    if pos + 2 > chunk_end:
                        pos = chunk_end
                        break
                    token = struct.unpack_from("<H", data, pos)[0]
                    pos += 2
                    diff = max(1, len(out) - chunk_start_out)
                    bit_count = max(4, int(math.ceil(math.log(diff, 2))))
                    length_mask = 0xFFFF >> bit_count
                    offset_mask = (~length_mask) & 0xFFFF
                    length = (token & length_mask) + 3
                    offset = ((token & offset_mask) >> (16 - bit_count)) + 1
                    copy_src = len(out) - offset
                    for i in range(length):
                        out.append(out[copy_src + i])
        pos = chunk_end
    return bytes(out)


def decode_vba_text(data: bytes) -> str:
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            if "\x00" not in text[:100]:
                return text
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="ignore")


def parse_vba_dir(dir_data: bytes) -> list[dict[str, Any]]:
    modules: list[dict[str, Any]] = []
    pos = 0
    module_count = None
    codepage = "cp1252"

    while pos + 6 <= len(dir_data):
        rec_id, rec_size = struct.unpack_from("<HI", dir_data, pos)
        pos += 6
        rec_data = dir_data[pos : pos + rec_size]
        pos += rec_size
        if rec_id == 0x0009 and pos + 2 <= len(dir_data):
            pos += 2
        if rec_id == 0x0003 and len(rec_data) >= 2:
            cp = struct.unpack_from("<H", rec_data, 0)[0]
            codepage = "cp1252" if cp in (0, 1252) else f"cp{cp}"
        if rec_id == DIR_RECORD_MODULES and len(rec_data) >= 2:
            module_count = struct.unpack_from("<H", rec_data, 0)[0]
            break

    if module_count is None:
        return modules

    def decode_name(raw: bytes) -> str:
        try:
            return raw.decode(codepage, errors="replace")
        except LookupError:
            return raw.decode("cp1252", errors="replace")

    for _ in range(module_count):
        module: dict[str, Any] = {}
        while pos + 6 <= len(dir_data):
            rec_id, rec_size = struct.unpack_from("<HI", dir_data, pos)
            pos += 6
            rec_data = dir_data[pos : pos + rec_size]
            pos += rec_size
            if rec_id == DIR_RECORD_MODULENAME:
                module["name"] = decode_name(rec_data)
            elif rec_id == DIR_RECORD_MODULESTREAMNAME:
                module["stream"] = decode_name(rec_data)
            elif rec_id == DIR_RECORD_MODULEDOCSTRING:
                module["docstring_length"] = rec_size
            elif rec_id == DIR_RECORD_MODULEOFFSET and len(rec_data) >= 4:
                module["offset"] = struct.unpack_from("<I", rec_data, 0)[0]
            elif rec_id == DIR_RECORD_MODULETYPE_PROCEDURAL:
                module["module_type"] = "procedural"
            elif rec_id == DIR_RECORD_MODULETYPE_DOCUMENT:
                module["module_type"] = "document"
            elif rec_id == DIR_RECORD_TERMINATOR:
                break
        if module:
            modules.append(module)

    return modules


def macro_profile(xlsm_path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(xlsm_path) as zf:
        if "xl/vbaProject.bin" not in zf.namelist():
            return {"present": False}
        binary = zf.read("xl/vbaProject.bin")
    result: dict[str, Any] = {"present": True, "modules": [], "references": [], "stream_count": 0}
    try:
        cfb = CompoundFile(binary)
        paths = cfb.list_paths()
        result["stream_count"] = len(paths)
        project_entry = paths.get("PROJECT")
        if project_entry:
            project_text = decode_vba_text(cfb.stream(project_entry))
            result["references"] = [
                line.split("=", 1)[1][:180]
                for line in project_text.splitlines()
                if line.startswith("Reference=")
            ][:20]

        dir_entry = paths.get("VBA/dir")
        dir_modules = parse_vba_dir(decompress_vba(cfb.stream(dir_entry))) if dir_entry else []
        result["dir_modules"] = dir_modules
        candidate_modules = dir_modules or [
            {"stream": path.split("/")[-1], "offset": 0, "module_type": None}
            for path, entry in sorted(paths.items())
            if entry.type == 2
            and path.startswith("VBA/")
            and path.split("/")[-1].lower() not in {"dir", "_vba_project"}
        ]

        for module in candidate_modules:
            leaf = module.get("stream")
            entry = paths.get(f"VBA/{leaf}")
            if not leaf or not entry:
                continue
            raw = cfb.stream(entry)
            offset = int(module.get("offset") or 0)
            decompressed = decompress_vba(raw[offset:])
            text = decode_vba_text(decompressed)
            procedures = [{"type": kind, "name": name} for kind, name in PROC_RE.findall(text)]
            if not procedures and not re.search(r"\b(Sub|Function|Property)\b", text, re.I):
                continue
            keyword_counts = {kw: len(re.findall(re.escape(kw), text, re.I)) for kw in CALL_KEYWORDS}
            result["modules"].append(
                {
                    "name": module.get("name") or leaf,
                    "stream": leaf,
                    "module_type": module.get("module_type"),
                    "line_count": len([line for line in text.splitlines() if line.strip()]),
                    "procedure_count": len(procedures),
                    "procedures": procedures,
                    "keyword_counts": {k: v for k, v in keyword_counts.items() if v},
                }
            )
    except Exception as exc:
        result["parse_error"] = str(exc)
    return result


def workbook_profile(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
    sheets = [sheet_profile(ws) for ws in wb.worksheets]
    defined_names = []
    try:
        for name in wb.defined_names:
            destination = str(name.attr_text) if getattr(name, "attr_text", None) else None
            defined_names.append(
                {
                    "name": name.name,
                    "scope": name.localSheetId,
                    "destination": re.sub(r"\$?[A-Z]{1,3}\$?\d+", "CELL", destination or "")[:180],
                }
            )
    except Exception:
        defined_names = []

    return {
        "file": path.name,
        "file_size_bytes": path.stat().st_size,
        "sheet_count": len(wb.worksheets),
        "sheets": sheets,
        "defined_names": defined_names[:100],
        "defined_name_count": len(defined_names),
        "zip_features": xml_feature_profile(path),
        "macros": macro_profile(path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, default=Path("analysis/excel_audit_sanitized.json"))
    args = parser.parse_args()

    profile = workbook_profile(args.workbook)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(args.out),
        "sheet_count": profile["sheet_count"],
        "sheets": [
            {
                "name": sheet["name"],
                "rows": sheet["dimensions"]["max_row"],
                "columns": sheet["dimensions"]["max_column"],
                "formulas": sheet["formula_count"],
            }
            for sheet in profile["sheets"]
        ],
        "macros": {
            "present": profile["macros"].get("present"),
            "module_count": len(profile["macros"].get("modules", [])),
            "procedure_count": sum(m.get("procedure_count", 0) for m in profile["macros"].get("modules", [])),
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
